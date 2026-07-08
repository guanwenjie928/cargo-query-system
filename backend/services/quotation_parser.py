"""
货盘查询系统 - 报价表解析器
解析纵向产品块格式的报价 Excel 文件

每个产品占 ~8 行，字段名在 B 列，值在 C 列
品类 = Sheet 名（软胶/功能软胶产品/飞机杯/阳具与水晶套）
"""
import openpyxl
from typing import Dict, List, Any
from backend.config import CATEGORY_SHEETS
from backend.services.image_extractor import extract_images_from_sheet


def parse_category_sheet(ws, category: str) -> tuple:
    """
    解析品类 Sheet（纵向产品块格式）

    逻辑：
    1. 在 B 列扫描所有 "Model" 行，确定每个产品块的起始位置
    2. 从 Model 行向下扫描，直到下一个 Model 行
    3. 按 B 列字段名提取 C 列的值
    4. 同时提取 A 列（状态/唛头）、G 列（限价/包装/特征码）、H 列（头程费）

    返回：(产品列表, {SKU: 图片二进制数据})
    """
    # 找到所有 Model 行的位置（B列值为 "Model"）
    model_rows = []
    for row_idx in range(1, ws.max_row + 1):
        b_value = ws.cell(row=row_idx, column=2).value
        if b_value and str(b_value).strip() == "Model":
            model_rows.append(row_idx)

    products = []

    for i, start_row in enumerate(model_rows):
        end_row = model_rows[i + 1] if i + 1 < len(model_rows) else ws.max_row + 1

        product = {
            "sku": None,
            "product_name": None,
            "product_name_en": None,
            "weight": None,
            "price_rmb": None,
            "price_usd": None,
            "material": None,
            "hs_code": None,
            "limit_price": None,
            "packaging": None,
            "feature_code": None,
            "first_leg_fee": None,
            "product_status": None,
            "shipping_mark": None,
            "category": category,
        }

        description_count = 0  # 区分第一个Description(中文名)和第二个(英文品类)

        for row in range(start_row, end_row):
            a_value = ws.cell(row=row, column=1).value  # A列：状态/唛头
            b_value = ws.cell(row=row, column=2).value  # B列：字段名
            c_value = ws.cell(row=row, column=3).value  # C列：字段值
            g_value = ws.cell(row=row, column=7).value  # G列：限价/包装/特征码
            h_value = ws.cell(row=row, column=8).value  # H列：头程费

            # B列字段解析
            if b_value:
                field = str(b_value).strip()

                if field == "Model":
                    product["sku"] = str(c_value).strip() if c_value else None

                elif field == "Description":
                    description_count += 1
                    if description_count == 1:
                        product["product_name"] = str(c_value).strip() if c_value else None
                    else:
                        product["product_name_en"] = str(c_value).strip() if c_value else None

                elif field == "Weight" or field == " Weight":
                    product["weight"] = str(c_value).strip() if c_value else None

                elif field == "RMB":
                    try:
                        product["price_rmb"] = float(c_value) if c_value else None
                    except (TypeError, ValueError):
                        product["price_rmb"] = None

                elif field == "USD":
                    try:
                        product["price_usd"] = float(c_value) if c_value else None
                    except (TypeError, ValueError):
                        product["price_usd"] = None

                elif field == "Material":
                    product["material"] = str(c_value).strip() if c_value else None

                elif field == "HS CODE":
                    hs = str(c_value).strip() if c_value else None
                    # 过滤掉 "HS CODE" 占位文字
                    if hs and hs.upper() != "HS CODE":
                        product["hs_code"] = hs

                elif field == "Quantity":
                    # 飞机杯有此字段，作为重量
                    if c_value is not None:
                        product["weight"] = f"{c_value}kg"

                elif field == "EUR":
                    # 部分产品有欧元价格，暂不处理
                    pass

            # A列：状态/唛头
            if a_value:
                a_str = str(a_value).strip()
                if "下架" in a_str or "包销" in a_str:
                    product["product_status"] = a_str
                elif a_str and a_str != "None":
                    # 唛头信息（如 "6276 1/668-6276 100/668"）
                    if not product["shipping_mark"]:
                        product["shipping_mark"] = a_str

            # G列：限价/包装方式/特征码
            if g_value:
                g_str = str(g_value).strip()
                if "限价" in g_str or "美金" in g_str:
                    product["limit_price"] = g_str
                elif "包装" in g_str or "袋" in g_str or "盒" in g_str or "收缩膜" in g_str:
                    if not product["packaging"]:
                        product["packaging"] = g_str
                elif "特征码" in g_str:
                    if not product["feature_code"]:
                        product["feature_code"] = g_str
                    else:
                        product["feature_code"] = f"{product['feature_code']} / {g_str}"

            # H列：头程费
            if h_value is not None and isinstance(h_value, (int, float)):
                product["first_leg_fee"] = float(h_value)

        # 只保留有 SKU 的产品
        if product["sku"]:
            products.append(product)

    # 提取图片（图片锚定在 A 列附近，需要关联到产品块）
    images = extract_images_from_sheet(ws, sku_col=3, model_rows=model_rows)

    return products, images


def parse_fee_sheet(ws, region: str) -> List[Dict]:
    """解析订单处理费 Sheet（美东/美西）"""
    fees = []

    if region == "east":
        # 美东格式：SKU, lb, kg, 处理费（第1行是标题，第2行是列名）
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
            if not row or not row[0]:
                continue
            try:
                # 跳过非数据行（说明文字等），要求处理费必须有值
                fee = float(row[3]) if row[3] else None
                if fee is None:
                    continue
                fees.append({
                    "sku": str(row[0]).strip(),
                    "region": "east",
                    "weight_lb": float(row[1]) if row[1] else None,
                    "weight_kg": float(row[2]) if row[2] else None,
                    "processing_fee": fee,
                })
            except (TypeError, ValueError, IndexError):
                continue
    else:
        # 美西格式：仓库, SKU, 产品名称, 标签值, SKU核重(g), 核算尺寸(cm), 订单处理费$
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            if not row or not row[1]:
                continue
            try:
                fees.append({
                    "sku": str(row[1]).strip(),
                    "region": "west",
                    "warehouse": str(row[0]).strip() if row[0] else None,
                    "label_value": str(row[3]).strip() if row[3] else None,
                    "weight_g": int(row[4]) if row[4] else None,
                    "dimensions": str(row[5]).strip() if row[5] else None,
                    "processing_fee": float(row[6]) if row[6] else None,
                })
            except (TypeError, ValueError, IndexError):
                continue

    return fees


def parse_quotation_excel(file_path: str) -> Dict[str, Any]:
    """
    解析报价表 Excel 文件

    返回：
    {
        "products": [...],       # 所有品类产品
        "images": {sku: bytes},  # 图片数据
        "fees_east": [...],      # 美东处理费
        "fees_west": [...],      # 美西处理费
        "summary": {...}         # 汇总
    }
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    result = {
        "products": [],
        "images": {},
        "fees_east": [],
        "fees_west": [],
        "summary": {},
    }

    all_images = {}

    # 解析品类 Sheet
    for sheet_name in CATEGORY_SHEETS:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            products, images = parse_category_sheet(ws, sheet_name)
            result["products"].extend(products)
            all_images.update(images)

    # 解析美东订单处理费
    if "美东订单处理费" in wb.sheetnames:
        ws = wb["美东订单处理费"]
        result["fees_east"] = parse_fee_sheet(ws, "east")

    # 解析美西订单处理费
    if "美西订单处理费" in wb.sheetnames:
        ws = wb["美西订单处理费"]
        result["fees_west"] = parse_fee_sheet(ws, "west")

    result["images"] = all_images
    result["summary"] = {
        "product_count": len(result["products"]),
        "image_count": len(all_images),
        "fees_east_count": len(result["fees_east"]),
        "fees_west_count": len(result["fees_west"]),
        "categories": {cat: 0 for cat in CATEGORY_SHEETS},
    }
    for p in result["products"]:
        cat = p.get("category")
        if cat in result["summary"]["categories"]:
            result["summary"]["categories"][cat] += 1

    wb.close()
    del wb
    import gc; gc.collect()
    return result
