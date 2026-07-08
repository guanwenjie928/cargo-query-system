"""
货盘查询系统 - 库存表解析器
解析标准表格格式的库存 Excel 文件

支持两个 Sheet：
- stock: 多仓库分行（每个SKU 4行：Multiple/CA01/NC01虚拟仓/XISE）
- Sheet2: 一行一个SKU，含内嵌图片
"""
import openpyxl
from typing import Dict, List, Tuple, Any
from backend.services.image_extractor import extract_images_from_sheet


def parse_stock_sheet(ws) -> List[Dict]:
    """
    解析 stock Sheet（多仓库分行格式）

    返回格式：
    [
        {
            "sku": "XS-MA50021",
            "product_name": "Chiquita Ass娇娃",
            "in_transit": 0,
            "stock_total": 98,
            "warehouses": [
                {"warehouse": "Multiple", "stock": 98, "in_transit": 0},
                {"warehouse": "CA01", "stock": 0, "in_transit": 0},
                {"warehouse": "NC01虚拟仓", "stock": 0, "in_transit": 0},
                {"warehouse": "XISE", "stock": 98, "in_transit": 0},
            ]
        },
        ...
    ]
    """
    products = {}  # 按SKU聚合

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if not row or not row[0]:
            continue

        sku = str(row[0]).strip()
        product_name = str(row[1]).strip() if row[1] else None
        in_transit = int(row[2]) if row[2] is not None else 0
        stock = int(row[3]) if row[3] is not None else 0
        warehouse = str(row[4]).strip() if row[4] else ""

        if sku not in products:
            products[sku] = {
                "sku": sku,
                "product_name": product_name,
                "in_transit": 0,
                "stock_total": 0,
                "warehouses": [],
            }

        # 产品名只取有值的行
        if product_name and not products[sku]["product_name"]:
            products[sku]["product_name"] = product_name

        # Multiple 行的库存作为总库存
        if warehouse == "Multiple":
            products[sku]["stock_total"] = stock
            products[sku]["in_transit"] = in_transit

        # 记录各仓库明细
        products[sku]["warehouses"].append({
            "warehouse": warehouse,
            "stock": stock,
            "in_transit": in_transit,
        })

    return list(products.values())


def parse_sheet2(ws) -> Tuple[List[Dict], Dict[str, bytes]]:
    """
    解析 Sheet2（一行一个SKU + 内嵌图片）

    返回：(产品列表, {SKU: 图片二进制数据})
    """
    products = []
    sku_row_map = {}  # SKU -> Excel行号，用于图片关联

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), 2):
        if not row or not row[0]:
            continue

        sku = str(row[0]).strip()
        product_name = str(row[1]).strip() if row[1] else None
        in_transit = int(row[2]) if row[2] is not None else 0
        stock = int(row[3]) if row[3] is not None else 0

        products.append({
            "sku": sku,
            "product_name": product_name,
            "in_transit": in_transit,
            "stock_total": stock,
        })
        sku_row_map[sku] = row_idx

    # 提取内嵌图片
    images = extract_images_from_sheet(ws, sku_col=1)  # SKU在A列

    # 按行号关联SKU
    sku_images = {}
    for row_num, img_data in images.items():
        # images 的 key 是 1-based 行号
        # 找到该行对应的SKU
        for sku, excel_row in sku_row_map.items():
            if excel_row == row_num:
                sku_images[sku] = img_data
                break

    return products, sku_images


def parse_stock_excel(file_path: str) -> Dict[str, Any]:
    """
    解析库存表 Excel 文件

    返回：
    {
        "stock_products": [...],      # stock Sheet 解析结果
        "sheet2_products": [...],     # Sheet2 解析结果
        "images": {sku: bytes},       # 图片数据
        "summary": {...}              # 汇总信息
    }
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    result = {
        "stock_products": [],
        "sheet2_products": [],
        "images": {},
        "summary": {},
    }

    # 解析 stock Sheet
    if "stock" in wb.sheetnames:
        ws = wb["stock"]
        result["stock_products"] = parse_stock_sheet(ws)

    # 解析 Sheet2（含图片）
    if "Sheet2" in wb.sheetnames:
        ws = wb["Sheet2"]
        products, images = parse_sheet2(ws)
        result["sheet2_products"] = products
        result["images"] = images

    result["summary"] = {
        "stock_sheet_count": len(result["stock_products"]),
        "sheet2_count": len(result["sheet2_products"]),
        "image_count": len(result["images"]),
    }

    wb.close()
    return result
